# ensure that docker container jmsdb is running to be able to access the postgress database
# Running this app will start the flask server
# This API will qeury, add, save and delete data from the database.
# The query routine will return data in json format using the exact column name in the
# database, excluding the lock column. Its up to the UI to change the order of data and to
# decide which columns to use.

import traceback
from flask import Flask, jsonify, request
import psycopg2
from flask_cors import CORS
from PyPDF2 import PdfFileWriter, PdfFileReader, PdfFileMerger
# from os import remove, path, startfile, getcwd, makedirs
import os
from reportlab.pdfgen import canvas
import textwrap
import math
import time
import openpyxl

totalamount = 0

pdffolder = '../data/pdfs'
xldatafolder = '../data/xldata'
templatefolder = '../data/templates'
pdfoverlayfile = pdffolder + '/overlay.pdf'


app = Flask(__name__)
CORS(app)

# Database connection parameters
db_params = {
    'host': 'circle-rhino-postgres',
    # 'host': 'localhost',
    'dbname': 'jmsdb',
    'user': 'postgres',
    'password': 'postgres',
    'port': '5432'
    # 'port': '5430'
}
  
# Function to connect to the database
def connect_db():
    connection = psycopg2.connect(**db_params)
    return connection

@app.route('/hw', methods=['PATCH'])
def hello_world():
    print("HW endpoit was hit")
    try:
        connection = connect_db()
        cursor = connection.cursor()
        cursor.execute("""
        SELECT * FROM public.alembic_version
        """)
        alembic_version = cursor.fetchall()
        connection.close()
        return jsonify({'message': 'Hello World '}), 200
    except Exception as e:
          traceback.print_exc(e.__context__)
          return jsonify({'error': str(e)}), 500

# Route to get query data (GET - read). UI sent payload: (fieldnames, fieldaliases, joinfields)
@app.route('/query', methods=['POST'])
def get_data():
    try:
        thedata = request.json
        orderby = ''
        if 'query' in thedata.keys(): # use query code as is
            querycode = thedata['query']
            while '  ' in querycode:
                querycode = querycode.replace('  ', ' ')
        else: # build the query code
            # construct the query command
            fldnames = '' # contains the fields to return (SELECT)
            maintable = '' # main table with fields (FROM)
            joincode = '' # the code used to join fields (JOIN jt1 ON pk1=fk1 JOIN jt2 ON pk2=fk2 ....)
            filtercode = '' # how the selection is filtered (WHERE)
            for item in thedata.keys():
                if item == 'fieldnames': # field names to query
                    for index, fldn in enumerate(thedata['fieldnames']):
                        myfield = fldn + ' as ' + thedata['fieldaliases'][index] + ', '
                        fldnames = fldnames + myfield
                        if maintable == '':
                            maintable = myfield.split('.')[0]
                            if "(" in maintable:
                                posit = maintable.find('(')
                                maintable = maintable[posit+1:]
                if item == 'joinfields': # tables and fields to do join
                    for join in thedata[item]: # itterate through join list
                        joincode = joincode + ' LEFT JOIN ' + join['jointable'] + ' ON ' +  join['primarykey'] + '=' + join['foreinkey']
                if item == 'filterfields':
                    for fltr in thedata[item]:
                        filtercode = filtercode + ' ' + fltr['field'] + '=' + f"""'{fltr['value']}'""" + ' AND'
            filtercode = filtercode + ' ' + maintable + '.lock=false'
            fldnames = fldnames[:-2] # remove end ', ' and ' = '
            querycode = 'SELECT ' + fldnames + ' FROM ' + maintable
            if len(joincode) > 0:
                querycode = querycode + joincode
            querycode = querycode + ' WHERE' + filtercode
            orderby = 'ORDER BY id ASC'

        connection = connect_db() # Connect to the database
        cursor = connection.cursor() # Create a cursor
        print('query code: ', f"""{querycode} {orderby}""", '\n')
        cursor.execute(f"""{querycode} {orderby}""") # Execute a query
        data = cursor.fetchall() # Fetch all rows
        colnames = [desc[0] for desc in cursor.description]
        cursor.close()# Close the cursor and connection
        connection.close()
        result = []
        for row in data:
            myrow = {}
            for index, coln in enumerate(colnames):
                myrow[coln] = row[index]
            result.append(myrow)
        return jsonify(result)
    except Exception as e:
          traceback.print_exc(e.__context__)
          return jsonify({'error': str(e)}), 500
    
# Route to add new record (POST - create). UI sent payload: (table name, datafield1, datafield2, ...) (no id field)
@app.route('/add', methods=['POST'])
def add_data():
    try:
        # Get data from the request (assuming a JSON payload)
        thedata = request.json
        mykeys = ''
        mydata = ''
        for item in thedata.keys():
            if item == 'table': # first itteration
                table = thedata[item]
            else:
                if thedata[item] != None: # do not attempt to add fields with type none/null
                    mykeys = mykeys + item + ', '
                    dataitem = thedata[item].replace("'", '`')
                    dataitem = dataitem.replace("\\", '')
                    mydata = mydata + "'" + dataitem + "', "
        mykeys = (mykeys + 'lock')
        mydata = (mydata + 'false')
        
        # Connect to the database
        connection = connect_db()
        # Create a cursor
        cursor = connection.cursor()
        # Execute a query
        print("adding to db command:",'INSERT INTO', table, '(', mykeys, ') VALUES (', mydata, ')', '\n')
        cursor.execute(
            f"""
            INSERT INTO {table} ({mykeys})
            VALUES ({mydata});
            """
        )
        connection.commit()
        # Close the cursor and connection
        cursor.close()
        connection.close()
        # return "success"
        return jsonify({'message': 'record added successfully'}), 200
    except Exception as e:
        traceback.print_exc(e.__context__)
        return jsonify({'error': str(e)}), 500
        
# Route to update record (PATCH - update). UI sent payload: (table name, id, datafield1, datafield2, ...)
@app.route('/save', methods=['PATCH'])
def save_data():
    try:
        # Get data from the request (assuming a JSON payload)
        thedata = request.json
        mysetdata = ''
        for item in thedata.keys():
            if item == 'table':
                table = thedata[item]
            elif item == 'id':
                myid = thedata[item]
            else:
                dataitem = thedata[item].replace("'", '`')
                dataitem = dataitem.replace("\\", '')
                mysetdata = mysetdata + item + "='" + dataitem + "', "
        mysetdata = (mysetdata + 'lock=false')
        

        # Connect to the database
        connection = connect_db()
        # Create a cursor
        cursor = connection.cursor()
        # Execute a query
        print('save command: UPDATE', table, ' SET ', mysetdata, 'WHERE id=', myid, '\n')
        cursor.execute(
            f"""
                UPDATE {table}
                SET {mysetdata}
                WHERE id={myid}
            """
        )
        connection.commit()
        # Close the cursor and connection
        cursor.close()
        connection.close()
        return jsonify({'message': 'record saved successfully'}), 200
    except Exception as e:
        return jsonify({'error': str(e)}), 500
        
# Route to delete a record (PATCH - delete - update lock). UI sent payload: (table name, id)
@app.route('/delete', methods=['PATCH'])
def delete_data():
    # only changing lock to true
    try:
        # Get data from the request (assuming a JSON payload)
        thedata = request.json
        table = thedata['table']
        myid = thedata['id']
        # Connect to the database
        connection = connect_db()
        # Create a cursor
        cursor = connection.cursor()
        # Execute a query
        cursor.execute(
            f"""
                UPDATE {table}
                SET lock=true
                where id={myid}
            """
        )
        connection.commit()
        # Close the cursor and connection
        cursor.close()
        connection.close()
        return jsonify({'message': 'record locked successfully'}), 200
    except Exception as e:
        return jsonify({'error': str(e)}), 500

# Route to print records. UI sent payload: (type, id)
# where type is: jobcard or quote
#   and id: is the records id
@app.route('/print', methods=['POST'])
def print_record():
     # creates a pdf for the user to print or send
    global totalamount
    try:
        thedata = request.json
        if thedata['type'] == "quotes": # get quotelines
            querycode = f"""SELECT quotes.quote_no, TO_CHAR(quotes.quote_date, 'dd/mm/yyyy'), customers.company_name,
                        contacts.contact_person, contacts.tel_no, contacts.email
                        FROM quotes
                        LEFT JOIN contacts ON contacts.id=quotes.contacts_id 
                        LEFT JOIN customers ON customers.id=quotes.customers_id 
                        WHERE quotes.lock=false AND quotes.id ='{thedata['id']}'"""
            connection = connect_db() # Connect to the database
            cursor = connection.cursor() # Create a cursor
            cursor.execute(f"""{querycode}""") # Execute a query
            data = cursor.fetchall() # Fetch all rows
            # colnames = [desc[0] for desc in cursor.description]
            cursor.close()# Close the cursor and connection
            connection.close()
            # retrieve data lines
            querycode = f"""SELECT quote_lines.quantity, quote_lines.name, quote_lines.description,
                        quote_lines.price 
                        FROM quote_lines 
                        WHERE quote_lines.lock=false AND quote_lines.quotes_id ='{thedata['id']}'"""
            connection = connect_db() # Connect to the database
            cursor = connection.cursor() # Create a cursor
            cursor.execute(f"""{querycode}""") # Execute a query
            linedata = cursor.fetchall() # Fetch all rows
            cursor.close()# Close the cursor and connection
            connection.close()
            # calculate no of pages
            totallines = 0
            pagelines = 0
            itemsperpage = []
            for idx, i in enumerate(linedata):
                ilines = textwrap.wrap(str(i[1] + ": " + i[2]), 56)
                if pagelines + len(ilines) + 1 > 22:
                    itemsperpage.append(idx-1)
                    pagelines = len(ilines) + 1
                else:
                    pagelines = pagelines + len(ilines) + 1
                totallines = totallines + len(ilines) + 1
            itemsperpage.append(idx+1)
            # 22 lines on a page
            noofpages = math.ceil(totallines / 22)
            totalamount = 0
            pagestart = 0
            mypages = []
            for p in range(noofpages-1): # do page type 1
                mypages.append(createpage('jmsquotep1.pdf',data,linedata[pagestart:itemsperpage[p]+1], p + 1, noofpages))
                pagestart = itemsperpage[p]+1
            mypages.append(createpage('jmsquotep2.pdf',data,linedata[pagestart:itemsperpage[-1]], noofpages, noofpages))
            # merge and rename files
            merger = PdfFileMerger()
            for filename in mypages:
                myfile = open(filename,'rb')
                input = PdfFileReader(myfile)
                merger.append((input))
                myfile.close()
            finfile = "pdfs/quote_" + str(data[0][0]).zfill(6) + "_" + str(data[0][2]) + ".pdf"
            merger.write(finfile)
            # delete page files
            for filename in mypages:
                remove(filename)
        
        fullfilename = path.abspath(finfile)
        time.sleep(2)
        window_title = finfile.split('/')[1]
        print("opening file:", window_title)
        try:
            startfile(fullfilename)
            returnstring = window_title + " file opened successfully"
        except:
            returnstring = "failed to open " + window_title
        return jsonify(returnstring), 200
        #return jsonify({'message': 'print successfully'}), 200
    except Exception as e:
          traceback.print_exc(e.__context__)
          return jsonify({'error': str(e)}), 500
   
# Route to export a table. UI sent payload: (table name)
@app.route('/export', methods=['PATCH'])
def export_data():
    try:
        thedata = request.json
        #print('data received:', thedata)
        table = thedata['table']
        querycode = 'SELECT * FROM ' + table  + " WHERE lock=false ORDER BY id"
        print('exporting data query code:', querycode)
        connection = connect_db() # Connect to the database
        cursor = connection.cursor() # Create a cursor
        cursor.execute(f"""{querycode}""") # Execute a query
        data = cursor.fetchall() # Fetch all rows
        colnames = [desc[0] for desc in cursor.description]
        cursor.close()# Close the cursor and connection
        connection.close()
        # create excel file and export data
        try:
            workbook = openpyxl.Workbook()
            sheet = workbook.active
            # add header row
            headlist = []
            for coln in colnames:
                if coln != 'lock':
                    headlist.append(coln)
            sheet.append(headlist)
            # add data rows
            for row in data:
                myrow = []
                for index, coln in enumerate(colnames):
                    if coln != 'lock':
                        myrow.append(row[index])
                sheet.append(myrow)
            # Save the workbook to a file
            currentfolder = getcwd() + '\\data\\xldata'
            # create folder if not exist
            isExist = path.exists(currentfolder)
            #print('does path exist:', isExist)
            if not isExist:
                makedirs(currentfolder)
            xlfilename = currentfolder + '\\' + table + '.xlsx'
            #print(xlfilename)
            workbook.save(xlfilename)
            returnmessage = 'Success! Data exported succesfully to file ' + xlfilename
        except:
            returnmessage = 'Error! Could not create export file: ' + xlfilename
        return jsonify({'message': returnmessage}), 200
    except Exception as e:
        traceback.print_exc(e.__context__)
        return jsonify({'error': str(e)}), 500

# Route to import a table. UI sent payload: (table name)
@app.route('/import', methods=['PATCH'])
def import_data():
    try:
        thedata = request.json
        table = thedata['table']
        # extract data from file
        currentfolder = getcwd() + '\\data\\xldata'
        xlfilename = currentfolder + '\\' + table + '.xlsx'
        returnmessage = 'Success! Data imported succesfully from file ' + xlfilename # this message will change if errors occur
        isExist = path.exists(xlfilename)
        if not isExist:
            returnmessage = 'Error! Could not find file to import: ' + xlfilename
        else: # file exist
            try: # retrieve data from file
                workbook = openpyxl.load_workbook(xlfilename)
                sheet = workbook.active
                mysetfields = ''
                headerow = ''
                mykeys = ''
                for i in range(1,sheet.max_column+1):
                    colname = sheet.cell(row=1,column=i).value
                    headerow = headerow + colname + ', '
                    mysetfields = mysetfields + colname + '=' +  'newdata.' + colname + ', '
                    if colname != 'id':
                        mykeys = mykeys + colname + ', '
                headerow = headerow[:-2]
                mysetfields = mysetfields[:-2]
                mykeys = mykeys + 'lock'
                mysetdata = '('
                mydata = '('
                newcount = 0
                for myrow in range(2, sheet.max_row+1): # itterate through rows
                    for i in range(1,sheet.max_column+1): # itterate through columns
                        myvalue = str(sheet.cell(row=myrow,column=i).value).replace("'", '`')
                        myvalue = myvalue.replace("\\", '')
                        #print(myvalue)
                        if i == 1: # the id
                            theid = myvalue
                        if theid == 'None': # add to new record list
                            if i != 1: # dont add id
                                if myvalue.replace('.','').replace('-','').isnumeric(): # numeric data
                                    mydata = mydata + myvalue + ", "
                                else:
                                    mydata = mydata + "'" + myvalue + "', "
                        else: # add to save list
                            if myvalue.replace('.','').replace('-','').isnumeric(): # numeric data
                                mysetdata = mysetdata + myvalue + ", "
                            else:
                                mysetdata = mysetdata + "'" + myvalue + "', "
                    if theid == 'None': # add to new record list
                        mydata = mydata + 'false), ('
                        newcount = newcount + 1
                    else: # add to existing records to update
                        mysetdata = mysetdata[:-2] + '), ('
                mysetdata = mysetdata[:-3]
                mydata = mydata[:-3]
                # remove any fields that contains a value of 'None'
                mysetdata = mysetdata.replace("'None'", 'NULL')
                mydata = mydata.replace("'None'", 'NULL')

                # data extracted so store to db
                connection = connect_db() # Connect to the database
                cursor = connection.cursor() # Create a cursor
                try: # save existing records
                    print('SAVE command: UPDATE ' + table + ' AS mytable \nSET ' + mysetfields +
                            '\nFROM (VALUES ' + mysetdata + ') AS newdata(' + headerow +
                            ') \nWHERE newdata.id=mytable.id')
                    if len(mysetdata) > 0:
                        cursor.execute(
                            f"""
                                UPDATE {table} AS mytable
                                SET {mysetfields}
                                FROM (VALUES {mysetdata}) AS newdata({headerow})
                                WHERE newdata.id = mytable.id
                            """)
                        connection.commit()
                except:
                    returnmessage = 'Error! Existing data could not be updated in table ' + table
                try: # add new records
                    print("INSERT command:\n" +
                        '    INSERT INTO ' + table + ' ('+ mykeys + ')\n' +
                        '    VALUES '+ mydata + '\n')
                    if len(mydata) > 0:
                        cursor.execute(
                            f"""
                            INSERT INTO {table} ({mykeys})
                            VALUES {mydata}
                            """
                        )
                        connection.commit()
                except:
                    returnmessage = 'Error! New data could not be added to table ' + table
                # Close the cursor and connection
                cursor.close()
                connection.close()
            except:
                returnmessage = 'Error! Data could not be extracted from file ' + xlfilename
        return jsonify({'message': returnmessage}), 200
    except Exception as e:
        traceback.print_exc(e.__context__)
        return jsonify({'error': str(e)}), 500


# execute postgress commands as is in payload
@app.route('/action', methods=['PATCH'])
def action_data():
    # only changing lock to true
    try:
        # Get data from the request (assuming a JSON payload)
        thedata = request.json
        actioncode = thedata['query']
        # Connect to the database
        connection = connect_db()
        # Create a cursor
        cursor = connection.cursor()
        # Execute a query
        print('action code:', actioncode)
        cursor.execute(f"""{actioncode}""")
        connection.commit()
        # Close the cursor and connection
        cursor.close()
        connection.close()
        return jsonify({'message': 'action executed successfully'}), 200
    except Exception as e:
        return jsonify({'error': str(e)}), 500
    

def mergepdfs(template, outputfile): # where template is the template and filename is the outputfile
    finalout = PdfFileWriter()
    thetemplate = PdfFileReader(open(template, "rb"))
    overlay = PdfFileReader(open(pdfoverlayfile, "rb"))
    input_page = thetemplate.getPage(0)
    input_page.mergePage(overlay.getPage(0))
    finalout.addPage(input_page)
    # finally, write "output" to document-output.pdf
    outputStream = open(outputfile, "wb")
    finalout.write(outputStream)
    outputStream.close()

    
def createpage(pgtype, pgheader, pgdata, pgno, totpges): # creates a pdf page containing the data where:
    # pgtype is the template page name
    #Page layout: starting at bottom left, x to right, y tot top where 1mm = 2.8346 points
    # page size = 210mm(x) x 297mm(y)
    # quote area size = 178mm x129mm
    global totalamount
    mm = 2.8346
    #ensure no overlay file exist
    try:
        remove(pdfoverlayfile)
    except:
        pass
    #Create overlay
    mytempl = templatefolder + "/" + pgtype
    c = canvas.Canvas(pdfoverlayfile)
    c.setLineWidth(0.5)
    c.setFont('Times-Bold', 10)

    if pgtype == "jmsquotep1.pdf" or pgtype == "jmsquotep2.pdf": # 1st page of the quote
        filename = pdffolder + "/quote_" + str(pgheader[0][0]).zfill(6) + "_" + str(pgno) + ".pdf"
        # populate header data in quote
        # do page numbers
        corx =  158.0 * mm  # converted to canvas points
        cory =  (297 - 47.5) * mm # converted to canvas points
        c.drawString(corx,cory,str(pgno))
        corx =  166.0 * mm  # converted to canvas points
        c.drawString(corx,cory,str(totpges))
        #              quote no  , quote date ,  client   , attention ,  tel no   ,   email
        headerloc = [[167.8,42.2],[109.7,42.2],[27.7,42.2],[32.9,47.5],[23.8,52.4],[27.7,57.2]] # mm from top left
        for i, item in enumerate(pgheader[0]):
            corx =  headerloc[i][0] * mm  # converted to canvas points
            cory =  (297 - headerloc[i][1]) * mm # converted to canvas points
            if i == 0: # quote number
                c.drawString(corx,cory,str(item).zfill(6))
            else:
                c.drawString(corx,cory,str(item))
        # create quote lines
        posqty = 23 # qty x poaition
        posdescr = 34.4 # description x poaition
        posprice = 151 # price per x poaition
        postotprice = 171 # total price x poaition
        posline = 75 # divider lines x poaition
        lineheight = 5.6 # in mm
        deschars = 50 # the number of characters allowed in the description
        for quoteline in pgdata: # itterate through quotelines
            for i, item in enumerate(quoteline): # iterate trough items in quoteline
                cory =  (297 - posline) * mm # converted to canvas points
                if i == 0: # qty            
                    corx =  posqty * mm  # converted to canvas points
                    c.drawString(corx,cory,str(item))
                if i == 2: # description
                    corx =  posdescr * mm  # converted to canvas points
                    descrfull = textwrap.wrap(str(quoteline[i-1]) + ': ' + str(item), deschars)
                    nextline = posline
                    for phrase in descrfull:
                        cory =  (297 - nextline) * mm # converted to canvas points
                        c.drawString(corx,cory,phrase)
                        nextline = nextline + lineheight # a line is 6.9mm high
                if i == 3: # price
                    corx =  posprice * mm  # converted to canvas points
                    c.drawString(corx,cory,str(item))
                    corx =  postotprice * mm  # converted to canvas points
                    totalamount = totalamount + item * quoteline[0]
                    c.drawString(corx,cory,"{:.2f}".format(item * quoteline[0]))
            c.line(15.8 * mm, (297 - nextline) * mm, 194.8 * mm, (297 - nextline) * mm)
            posline = nextline + lineheight

        if pgtype == "jmsquotep2.pdf": # last page of the quote  
            # add the totals at bottom of page
            c.drawString(postotprice * mm, (297 - 204.2) * mm,"{:.2f}".format(totalamount)) # total amount
            c.drawString(postotprice * mm, (297 - (204.2 + lineheight)) * mm,"{:.2f}".format(float(totalamount) * 0.15)) # 15% vat
            c.drawString(postotprice * mm, (297 - (204.2 + 2 * lineheight)) * mm,"{:.2f}".format(float(totalamount) * 1.15)) # total

    c.showPage()
    c.save()
    mergepdfs(mytempl, filename)
    #startfile(filename)
    return filename
    

if __name__ == '__main__':
    app.run(debug=True)





